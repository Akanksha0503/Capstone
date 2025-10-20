import openpyxl

def read_keywords(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    steps = []

    for row in range(2, sheet.max_row + 1):
        steps.append({
            "Keyword": sheet.cell(row, 2).value,
            "LocatorType": sheet.cell(row, 3).value,
            "LocatorValue": sheet.cell(row, 4).value,
            "TestData": sheet.cell(row, 5).value
        })

    return steps
