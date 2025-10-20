import os
import csv
import xml.etree.ElementTree as ET
import openpyxl


def load_test_data(file_path):
    """
    Load test data from CSV, Excel (.xlsx/.xls), or XML file.
    Returns a list of dictionaries for use in @pytest.mark.parametrize.
    """
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return _read_csv(file_path)
    elif ext in [".xls", ".xlsx"]:
        return _read_excel(file_path)
    elif ext == ".xml":
        return _read_xml(file_path)
    else:
        raise ValueError(f"Unsupported file format: {ext}")


# -------------------------------------------------------------------
# CSV Loader
# -------------------------------------------------------------------
import csv

def _read_csv(file_path):
    """
    Reads a CSV file with a single 'role' column.
    Returns a list of dictionaries like [{'role': 'Registered'}, {'role': 'Administrators'}, ...]
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"CSV file not found: {file_path}")

    data = []
    with open(file_path, newline="", encoding="utf-8") as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            clean_row = {k.strip(): v.strip() for k, v in row.items() if v}  # remove empty values
            data.append(clean_row)

    return data



# -------------------------------------------------------------------
# Excel Loader
# -------------------------------------------------------------------
def _read_excel(file_path):
    """Read data from Excel file and return list of dictionaries."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    headers = [cell.value.strip().lower() if isinstance(cell.value, str) else str(cell.value).lower()
               for cell in next(sheet.iter_rows(min_row=1, max_row=1))]

    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if any(row):  # Skip completely empty rows
            clean_row = {headers[i]: (str(row[i]).strip() if row[i] else "") for i in range(len(headers))}
            data.append(clean_row)
    return data


# -------------------------------------------------------------------
# XML Loader
# -------------------------------------------------------------------
def _read_xml(file_path):
    """Read data from XML file and return list of dictionaries."""
    tree = ET.parse(file_path)
    root = tree.getroot()

    data = []
    for child in root:
        entry = {}
        for element in child:
            key = element.tag.strip().lower()
            value = element.text.strip() if element.text else ""
            entry[key] = value
        if any(entry.values()):
            data.append(entry)
    return data
